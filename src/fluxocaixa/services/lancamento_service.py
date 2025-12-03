from ..domain import LancamentoCreate, LancamentoOut
from ..repositories import LancamentoRepository
import csv
import openpyxl
from io import BytesIO, StringIO
from datetime import date, datetime
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..models import (
    db,
    Lancamento,
    Qualificador,
    TipoLancamento,
    OrigemLancamento,
    ContaBancaria,
    Conferencia,
)


def list_lancamentos(
    start_date: date | None = None,
    end_date: date | None = None,
    tipo: int | None = None,
    qualificador_folha: int | None = None,
    seq_conta: int | None = None,
    cod_origem: int | None = None,
    page: int = 1,
    per_page: int = 50,
    sort_by: str = 'dat_lancamento',
    sort_order: str = 'desc',
    repo: LancamentoRepository | None = None
) -> tuple[list, int]:
    repo = repo or LancamentoRepository()
    return repo.list(
        start_date=start_date,
        end_date=end_date,
        tipo=tipo,
        qualificador_folha=qualificador_folha,
        seq_conta=seq_conta,
        cod_origem=cod_origem,
        page=page,
        per_page=per_page,
        sort_by=sort_by,
        sort_order=sort_order
    )


def list_tipos_lancamento():
    from ..repositories import TipoLancamentoRepository
    repo = TipoLancamentoRepository()
    return repo.list_all()


def list_origens_lancamento():
    from ..repositories import OrigemLancamentoRepository
    repo = OrigemLancamentoRepository()
    return repo.list_all()


def list_contas_bancarias():
    from ..repositories import ContaBancariaRepository
    repo = ContaBancariaRepository()
    return repo.list_active()


def list_conferencias():
    from ..repositories import ConferenciaRepository
    repo = ConferenciaRepository()
    return repo.list_all()


def create_lancamento(data: LancamentoCreate, repo: LancamentoRepository | None = None) -> LancamentoOut:
    repo = repo or LancamentoRepository()
    lanc = repo.create(data)
    return LancamentoOut(
        seq_lancamento=lanc.seq_lancamento,
        dat_lancamento=lanc.dat_lancamento,
        seq_qualificador=lanc.seq_qualificador,
        val_lancamento=lanc.val_lancamento,
        cod_tipo_lancamento=lanc.cod_tipo_lancamento,
        cod_origem_lancamento=lanc.cod_origem_lancamento,
        dsc_lancamento=None,
    seq_conta=lanc.seq_conta,
    )


def update_lancamento(ident: int, data: LancamentoCreate, repo: LancamentoRepository | None = None) -> LancamentoOut:
    repo = repo or LancamentoRepository()
    lanc = repo.update(ident, data)
    return LancamentoOut(
        seq_lancamento=lanc.seq_lancamento,
        dat_lancamento=lanc.dat_lancamento,
        seq_qualificador=lanc.seq_qualificador,
        val_lancamento=lanc.val_lancamento,
        cod_tipo_lancamento=lanc.cod_tipo_lancamento,
        cod_origem_lancamento=lanc.cod_origem_lancamento,
        dsc_lancamento=None,
    seq_conta=lanc.seq_conta,
    )


def delete_lancamento(ident: int, repo: LancamentoRepository | None = None):
    repo = repo or LancamentoRepository()
    repo.soft_delete(ident)


def import_lancamentos_service(
    content: bytes, filename: str, session: Session | None = None
) -> dict:
    session = session or db.session
    rows: list[dict] = []
    
    if filename.lower().endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(StringIO(text))
        for row in reader:
            rows.append({k.strip(): v for k, v in row.items()})
    elif filename.lower().endswith(('.xlsx', '.xls')):
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c).strip() if c else '' for c in next(ws.iter_rows(values_only=True))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            rows.append(data)
    else:
        return {"sucesso": 0, "erros": ["Formato de arquivo não suportado"]}

    # Pre-fetch data to avoid N+1 queries
    all_qualificadores = session.query(Qualificador).all()
    qualificadores_map = {q.dsc_qualificador.lower(): q for q in all_qualificadores}

    all_tipos = session.query(TipoLancamento).all()
    tipos_map = {t.dsc_tipo_lancamento.lower(): t.cod_tipo_lancamento for t in all_tipos}

    all_origens = session.query(OrigemLancamento).all()
    origens_map = {o.dsc_origem_lancamento.lower(): o.cod_origem_lancamento for o in all_origens}
    
    # Use "Importado" origin for imported entries
    origem_importado = session.query(OrigemLancamento).filter_by(dsc_origem_lancamento='Importado').first()
    origem_importado_cod = origem_importado.cod_origem_lancamento if origem_importado else None

    all_contas = session.query(ContaBancaria).all()
    contas_map = {(c.cod_banco, c.num_agencia, c.num_conta): c for c in all_contas}

    count = 0
    errors = []

    def get_or_create_conta(banco, agencia, conta):
        if not (banco and agencia and conta):
            return None
        banco = str(banco).strip()
        agencia = str(agencia).strip()
        conta = str(conta).strip()
        
        key = (banco, agencia, conta)
        if key in contas_map:
            return contas_map[key]
            
        c = ContaBancaria(
            cod_banco=banco,
            num_agencia=agencia,
            num_conta=conta,
            dsc_conta=f"{banco}-{agencia}/{conta}",
        )
        session.add(c)
        session.flush()
        contas_map[key] = c
        return c


    try:
        for i, item in enumerate(rows, start=2):
            dat = item.get('Data') or item.get('dat_lancamento')
            desc = item.get('Qualificador') or item.get('Descrição') or item.get('descricao')
            valor = item.get('Valor (R$)') or item.get('val_lancamento')
            tipo_raw = item.get('Tipo') or item.get('cod_tipo_lancamento')

            if not (dat and desc and valor and tipo_raw):
                errors.append(f"Linha {i}: Dados incompletos (Data, Qualificador, Valor ou Tipo faltando)")
                continue

            if isinstance(dat, datetime):
                dat = dat.date()
            elif isinstance(dat, str):
                try:
                    dat = date.fromisoformat(dat)
                except ValueError:
                    errors.append(f"Linha {i}: Data inválida '{dat}'")
                    continue

            qual = qualificadores_map.get(str(desc).lower())
            if not qual:
                errors.append(f"Linha {i}: Qualificador não encontrado para '{desc}'")
                continue

            if isinstance(tipo_raw, str) and not tipo_raw.isdigit():
                tipo = tipos_map.get(tipo_raw.lower())
                if not tipo:
                    errors.append(f"Linha {i}: Tipo inválido '{tipo_raw}'")
                    continue
            else:
                tipo = int(tipo_raw)

            # Use Importado origin for all imported entries
            if not origem_importado_cod:
                errors.append(f"Linha {i}: Origem 'Importado' não encontrada no sistema")
                continue

            # Detect optional bank fields
            banco = item.get('Banco') or item.get('banco') or item.get('BANCO')
            agencia = item.get('Agencia') or item.get('agencia') or item.get('AGENCIA')
            conta = item.get('Conta') or item.get('conta') or item.get('CONTA')
            conta_obj = get_or_create_conta(banco, agencia, conta)

            lanc = Lancamento(
                dat_lancamento=dat,
                seq_qualificador=qual.seq_qualificador,
                val_lancamento=float(valor),
                cod_tipo_lancamento=tipo,
                cod_origem_lancamento=origem_importado_cod,
                cod_pessoa_inclusao=1,
                seq_conta=(conta_obj.seq_conta if conta_obj else None),
            )
            session.add(lanc)
            count += 1
        
        session.commit()
        return {"sucesso": count, "erros": errors}
    except Exception as e:
        session.rollback()
        errors.append(f"Erro fatal durante importação: {str(e)}")
        return {"sucesso": 0, "erros": errors}
