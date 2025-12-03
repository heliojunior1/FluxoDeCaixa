"""Service layer for SaldoConta (Bank Account Balance) operations."""
from datetime import date
from io import BytesIO, StringIO
import csv
import openpyxl

from ..domain import SaldoContaCreate, SaldoContaUpdate
from ..repositories import SaldoContaRepository, ContaBancariaRepository


def list_saldos_conta(
    seq_conta: int | None = None,
    data_inicio: date | None = None,
    data_fim: date | None = None,
    page: int = 1,
    per_page: int = 50,
    sort_by: str = 'dat_saldo',
    sort_order: str = 'desc',
    repo: SaldoContaRepository | None = None
) -> tuple[list, int]:
    """List bank account balances with optional filters and pagination.
    
    Args:
        seq_conta: Optional account filter
        data_inicio: Optional start date filter
        data_fim: Optional end date filter
        page: Page number (1-indexed)
        per_page: Items per page
        sort_by: Column to sort by
        sort_order: 'asc' or 'desc'
        repo: Optional repository instance for dependency injection
        
    Returns:
        Tuple of (List of SaldoConta objects, total count)
    """
    repo = repo or SaldoContaRepository()
    return repo.list(
        seq_conta=seq_conta,
        data_inicio=data_inicio,
        data_fim=data_fim,
        page=page,
        per_page=per_page,
        sort_by=sort_by,
        sort_order=sort_order
    )


def get_saldo_conta(seq_saldo_conta: int, repo: SaldoContaRepository | None = None):
    """Get a single bank account balance by ID.
    
    Args:
        seq_saldo_conta: Balance record ID
        repo: Optional repository instance
        
    Returns:
        SaldoConta object or None if not found
    """
    repo = repo or SaldoContaRepository()
    from ..models import SaldoConta
    return repo.session.query(SaldoConta).get(seq_saldo_conta)


def create_saldo_conta(data: SaldoContaCreate, repo: SaldoContaRepository | None = None) -> tuple[object | None, str | None]:
    """Create a new bank account balance record.
    
    Args:
        data: Validated creation data
        repo: Optional repository instance
        
    Returns:
        Tuple of (Created SaldoConta object, error message)
        If successful: (SaldoConta, None)
        If error: (None, error_message)
    """
    repo = repo or SaldoContaRepository()
    # Default user ID to 1 (system user)
    return repo.create(
        seq_conta=data.seq_conta,
        dat_saldo=data.dat_saldo,
        val_saldo=data.val_saldo,
        cod_pessoa_inclusao=1
    )


def update_saldo_conta(
    seq_saldo_conta: int,
    data: SaldoContaUpdate,
    repo: SaldoContaRepository | None = None
):
    """Update an existing bank account balance record.
    
    Args:
        seq_saldo_conta: Balance record ID
        data: Validated update data
        repo: Optional repository instance
        
    Returns:
        Updated SaldoConta object or None if not found
    """
    repo = repo or SaldoContaRepository()
    # Default user ID to 1 (system user)
    return repo.update(
        seq_saldo_conta=seq_saldo_conta,
        val_saldo=data.val_saldo,
        cod_pessoa_alteracao=1
    )


def delete_saldo_conta(seq_saldo_conta: int, repo: SaldoContaRepository | None = None):
    """Delete a bank account balance record.
    
    Args:
        seq_saldo_conta: Balance record ID
        repo: Optional repository instance
        
    Returns:
        True if deleted, False if not found
    """
    repo = repo or SaldoContaRepository()
    return repo.delete(seq_saldo_conta)


def import_saldos_service(content: bytes, filename: str) -> dict:
    """Import multiple bank account balance records from a CSV or XLSX file.
    
    Args:
        content: File content as bytes
        filename: Name of the uploaded file
        
    Returns:
        Dictionary with 'sucesso' count and 'erros' list
    """
    rows: list[dict] = []
    
    # Parse file based on extension
    if filename.lower().endswith('.csv'):
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(StringIO(text))
        for row in reader:
            rows.append({k.strip(): v for k, v in row.items()})
    elif filename.lower().endswith(('.xlsx', '.xls')):
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        headers = [str(c).strip() if c else '' for c in next(ws.iter_rows(values_only=True))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Skip empty rows or instruction rows
            if not row or row[0] is None or (isinstance(row[0], str) and row[0].startswith('INSTRUÇÕES')):
                continue
            data = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            rows.append(data)
    else:
        return {"sucesso": 0, "erros": ["Formato de arquivo não suportado"]}
    
    # Initialize repositories
    saldo_repo = SaldoContaRepository()
    conta_repo = ContaBancariaRepository()
    
    # Pre-fetch accounts
    contas = conta_repo.list_active()
    contas_map = {}
    for conta in contas:
        # Key format: "banco-agencia-conta" (lowercase, no dashes in conta)
        conta_num = str(conta.num_conta).replace('-', '').strip()
        key = f"{str(conta.cod_banco).strip()}-{str(conta.num_agencia).strip()}-{conta_num}".lower()
        contas_map[key] = conta.seq_conta
        # Also add with dash in conta number
        key2 = f"{str(conta.cod_banco).strip()}-{str(conta.num_agencia).strip()}-{str(conta.num_conta).strip()}".lower()
        contas_map[key2] = conta.seq_conta
    
    count = 0
    errors = []
    saldos_data = []
    
    try:
        for i, item in enumerate(rows, start=2):
            # Get fields with various naming conventions
            dat = item.get('Data') or item.get('data') or item.get('dat_saldo')
            
            # Check for combined "Conta" field first (format: Banco/Agência/Número)
            conta_combined = item.get('Conta') or item.get('conta')
            
            if conta_combined and '/' in str(conta_combined):
                # Parse combined format: "001/12345/123456-7"
                parts = str(conta_combined).split('/')
                if len(parts) >= 3:
                    banco = parts[0].strip()
                    agencia = parts[1].strip()
                    conta = parts[2].strip()
                else:
                    errors.append(f"Linha {i}: Formato de conta inválido '{conta_combined}' (esperado: Banco/Agência/Número)")
                    continue
            else:
                # Separate fields
                banco = item.get('Banco') or item.get('banco') or item.get('cod_banco')
                agencia = item.get('Agência') or item.get('Agencia') or item.get('agencia') or item.get('num_agencia')
                conta = conta_combined or item.get('num_conta')
            
            valor = item.get('Saldo') or item.get('saldo') or item.get('Valor') or item.get('valor') or item.get('val_saldo')
            
            # Validate required fields
            if not dat:
                errors.append(f"Linha {i}: Data não informada")
                continue
            if not (banco and agencia and conta):
                errors.append(f"Linha {i}: Dados de conta incompletos (Banco, Agência ou Conta faltando)")
                continue
            if valor is None or valor == '':
                errors.append(f"Linha {i}: Valor/Saldo não informado")
                continue
            
            # Parse date
            from datetime import datetime
            if isinstance(dat, datetime):
                dat = dat.date()
            elif isinstance(dat, date):
                pass  # Already a date
            elif isinstance(dat, str):
                try:
                    dat = date.fromisoformat(dat)
                except ValueError:
                    # Try other formats
                    try:
                        dat = datetime.strptime(dat, '%d/%m/%Y').date()
                    except ValueError:
                        errors.append(f"Linha {i}: Data inválida '{dat}'")
                        continue
            
            # Find account - try with and without dashes
            conta_num = str(conta).replace('-', '').strip()
            key = f"{str(banco).strip()}-{str(agencia).strip()}-{conta_num}".lower()
            seq_conta = contas_map.get(key)
            if not seq_conta:
                # Try with original format (with dash)
                key2 = f"{str(banco).strip()}-{str(agencia).strip()}-{str(conta).strip()}".lower()
                seq_conta = contas_map.get(key2)
            if not seq_conta:
                errors.append(f"Linha {i}: Conta não encontrada '{banco}/{agencia}/{conta}'")
                continue
            
            # Parse value
            try:
                if isinstance(valor, str):
                    # Handle comma as decimal separator
                    valor = valor.replace(',', '.')
                val_saldo = float(valor)
            except (ValueError, TypeError):
                errors.append(f"Linha {i}: Valor inválido '{valor}'")
                continue
            
            # Add to batch
            saldos_data.append({
                'seq_conta': seq_conta,
                'dat_saldo': dat,
                'val_saldo': val_saldo,
                'cod_pessoa_inclusao': 1,
                'dat_inclusao': date.today()
            })
        
        # Bulk insert
        created = 0
        skipped = 0
        if saldos_data:
            created, skipped, bulk_errors = saldo_repo.bulk_create(saldos_data)
            errors.extend(bulk_errors)
        
        if skipped > 0:
            errors.insert(0, f"{skipped} registro(s) ignorado(s) por já existirem no sistema")
        
        return {"sucesso": created, "erros": errors}
    except Exception as e:
        errors.append(f"Erro fatal durante importação: {str(e)}")
        return {"sucesso": 0, "erros": errors}
