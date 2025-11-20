from datetime import date
import calendar
from io import BytesIO

from fastapi import Request, UploadFile, File
from fastapi.responses import RedirectResponse, StreamingResponse
import openpyxl

from . import router, templates, handle_exceptions
from ..services import (
    list_cenarios,
    get_cenario,
    create_cenario_with_ajustes,
    update_cenario_with_ajustes,
    delete_cenario,
    get_ajustes_dict,
    list_receita_qualificadores,
    list_despesa_qualificadores,
    list_active_qualificadores,
    get_qualificador_by_name,
)

@router.get('/projecoes')
@handle_exceptions
async def projecoes_menu(request: Request):
    """Menu principal das projeções."""
    return templates.TemplateResponse('projecoes_menu.html', {'request': request})


@router.get('/projecoes/cenarios', name='projecoes_cenarios')
@handle_exceptions
async def projecoes_cenarios(request: Request):
    cenarios = list_cenarios()
    qualificadores_receita = list_receita_qualificadores()
    qualificadores_despesa = list_despesa_qualificadores()
    meses_nomes = {i: calendar.month_name[i].capitalize() for i in range(1, 13)}
    return templates.TemplateResponse(
        'cenarios.html',
        {
            'request': request,
            'cenarios': cenarios,
            'qualificadores_receita': qualificadores_receita,
            'qualificadores_despesa': qualificadores_despesa,
            'meses_nomes': meses_nomes,
            'current_year': date.today().year,
        },
    )


@router.get('/projecoes/modelos')
@handle_exceptions
async def projecoes_modelos(request: Request):
    """Tela de projeções automáticas."""
    return templates.TemplateResponse('projecoes_modelos.html', {'request': request})


@router.get('/projecoes/get/{id}')
@handle_exceptions
async def get_cenario_route(id: int):
    cenario = get_cenario(id)
    if not cenario:
        return {"error": "Cenario not found"}, 404
        
    ajustes_dict = get_ajustes_dict(id)
    return {
        'seq_cenario': cenario.seq_cenario,
        'nom_cenario': cenario.nom_cenario,
        'dsc_cenario': cenario.dsc_cenario,
        'ajustes': ajustes_dict,
    }


@router.get('/projecoes/template-xlsx')
@handle_exceptions
async def download_cenario_template():
    """Return an XLSX template with all qualificadores listed."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nome", "Periodo", "Valor", "Percentual"])
    qualificadores = [q for q in list_active_qualificadores() if q.cod_qualificador_pai is not None]
    
    for i, q in enumerate(qualificadores, start=1):
        month = (i % 12) + 1
        valor = None
        percentual = None
        if q.dsc_qualificador.upper() == "ICMS":
            month = 1
            valor = 1200
        elif i % 2 == 0:
            valor = i * 100.0
        else:
            percentual = (i % 5 + 1) * 1.0
        ws.append([
            q.dsc_qualificador,
            f"{month:02d}/2025",
            valor,
            percentual,
        ])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    headers = {
        'Content-Disposition': 'attachment; filename="cenario_template.xlsx"'
    }
    return StreamingResponse(
        stream,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers=headers,
    )


@router.post('/projecoes/import-xlsx')
@handle_exceptions
async def import_cenario_xlsx(file: UploadFile = File(...)):
    """Parse an uploaded XLSX file and return adjustments mapping."""
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    ajustes_dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        nome, periodo, valor, percentual = (list(row) + [None, None, None, None])[:4]
        if not nome or not periodo:
            continue
        
        qual = get_qualificador_by_name(str(nome))
        if not qual:
            continue
        # Parse period to year and month
        ano = None
        mes = None
        if hasattr(periodo, 'year') and hasattr(periodo, 'month'):
            ano = periodo.year
            mes = periodo.month
        else:
            pstr = str(periodo)
            if '/' in pstr:
                part = pstr.split('/')
                if len(part) == 2:
                    mes = int(part[0])
                    ano = int(part[1])
            elif '-' in pstr:
                part = pstr.split('-')
                if len(part) >= 2:
                    ano = int(part[0])
                    mes = int(part[1])
        if not ano or not mes:
            continue
        if percentual is not None:
            cod = 'P'
            val = float(percentual)
        elif valor is not None:
            cod = 'V'
            val = float(valor)
        else:
            continue
        key = f"{ano}_{mes}_{qual.seq_qualificador}"
        ajustes_dict[key] = {
            'ano': int(ano),
            'mes': int(mes),
            'cod_tipo_ajuste': cod,
            'val_ajuste': val,
        }
    return {'ajustes': ajustes_dict}


@router.post('/projecoes/add', name='add_cenario')
@handle_exceptions
async def add_cenario_route(request: Request):
    form = await request.form()
    nom_cenario = form.get('nom_cenario')
    dsc_cenario = form.get('dsc_cenario')
    ano = int(form.get('ano'))

    # Convert form data to dict for service
    ajustes_data = dict(form)
    
    create_cenario_with_ajustes(nom_cenario, dsc_cenario, ano, ajustes_data)
    
    return RedirectResponse(request.url_for('projecoes_cenarios'), status_code=303)


@router.get('/projecoes/edit/{id}')
@router.post('/projecoes/edit/{id}')
@handle_exceptions
async def edit_cenario_route(request: Request, id: int):
    cenario = get_cenario(id)
    if not cenario:
        # Handle not found
        return RedirectResponse(request.url_for('projecoes_cenarios'), status_code=303)

    if request.method == 'POST':
        form = await request.form()
        nom_cenario = form['nom_cenario']
        dsc_cenario = form.get('dsc_cenario')
        ano = int(form.get('ano'))
        
        ajustes_data = dict(form)
        
        update_cenario_with_ajustes(id, nom_cenario, dsc_cenario, ano, ajustes_data)
        
        return RedirectResponse(request.url_for('projecoes_cenarios'), status_code=303)
    
    qualificadores = list_active_qualificadores()
    ajustes = get_ajustes_dict(id)
    
    # Convert dict keys back to tuple for template compatibility if needed, 
    # but template seems to iterate or lookup. 
    # The original code passed a dict with tuple keys: {(ano, mes, seq): obj}
    # The service returns { "ano_mes_seq": dict }
    # We need to adapt to what the template expects.
    # Let's check the template usage.
    # Template `cenario_edit.html` likely uses `ajustes[(ano, mes, q.seq_qualificador)]`
    
    # Reconstruct the dict with tuple keys for the template
    ajustes_tuple_keys = {}
    for key_str, data in ajustes.items():
        # key_str is "ano_mes_seq"
        parts = key_str.split('_')
        if len(parts) == 3:
            k = (int(parts[0]), int(parts[1]), int(parts[2]))
            # The template expects an object with .cod_tipo_ajuste and .val_ajuste attributes
            # We can use a simple class or namedtuple, or just a dict if the template supports dot access on dicts (Jinja2 does not by default for dicts, it uses [])
            # But wait, the original code passed the SQLAlchemy object.
            # Let's create a dummy object
            class AjusteObj:
                def __init__(self, d):
                    self.cod_tipo_ajuste = d['cod_tipo_ajuste']
                    self.val_ajuste = d['val_ajuste']
            
            ajustes_tuple_keys[k] = AjusteObj(data)

    receitas = [q for q in qualificadores if q.tipo_fluxo == 'receita' and q.cod_qualificador_pai is not None]
    despesas = [q for q in qualificadores if q.tipo_fluxo == 'despesa' and q.cod_qualificador_pai is not None]
    
    # Determine ano from adjustments or default to current year
    ano = date.today().year
    if ajustes_tuple_keys:
        ano = list(ajustes_tuple_keys.keys())[0][0]
        
    meses_nomes = {i: calendar.month_name[i].capitalize() for i in range(1, 13)}
    return templates.TemplateResponse(
        'cenario_edit.html',
        {
            'request': request,
            'cenario': cenario,
            'ajustes': ajustes_tuple_keys,
            'receitas': receitas,
            'despesas': despesas,
            'ano': ano,
            'meses_nomes': meses_nomes,
        },
    )


@router.post('/projecoes/delete/{id}', name='delete_cenario')
@handle_exceptions
async def delete_cenario_route(request: Request, id: int):
    delete_cenario(id)
    return RedirectResponse(request.url_for('projecoes_cenarios'), status_code=303)
